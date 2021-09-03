from Memorizacion_crud import updates
from openpyxl import load_workbook
from openpyxl.utils.cell import rows_from_range
archivo_excel = './Guardar_cronograma.xlsx'
file = load_workbook(archivo_excel)
# Leer
def read():
    todos = []
    sheet = file['Hoja1']
    rows = list(sheet.rows)
    for row in rows[1:]:
        dictionary = {
            'id': row[0].value,
            'actividad': row[1].value,
            'Estado': row[2].value,
            'Finalizado': row[3].value,
        }
        todos.append(dictionary)
    print('_________________')
    return todos
#Agregar uno nuevo
def create_dato(dato_a_pasar: dict):
    sheet = file['Hoja1']
    row_idx = sheet.max_row+1
    print(row_idx)
    sheet.cell(column=1, row=row_idx).value = dato_a_pasar.get('Id')
    sheet.cell(column=2, row=row_idx).value = dato_a_pasar.get('actividad')
    sheet.cell(column=3, row=row_idx).value = dato_a_pasar.get('Estado')
    sheet.cell(column=4, row=row_idx).value = dato_a_pasar.get('Finalizado')
    file.save(archivo_excel)

#Eliminar archivo

def eliminar(id:int):
    shet=file['Hoja1']
    for x in shet.rows:
       if x[0].value==id:
           idx=x[0].row
           shet.delete_rows(idx)
opciones_usuario:dict={
'0':'Read the file excel',
'1':'Agregar un dato en el archivo excel', 
'2':'eliminar un dato especifico por el id',
'3':'Actualizar uno dato especifico'
}
print(opciones_usuario)
def actualizar_dato(id:int, data:dict):
    sheet=file['Hoja1']
    for x in sheet.rows:
       if x[0].value==id:
          idx=x[0].row
          if data.get('actividad'):sheet.cell(row=idx, column=1).value=data.get('actividad')
          if data.get('Estado'):sheet.cell(column=2, row=idx).value=data.get('Estado')
          if data.get('Finalizado'):sheet.cell(row=idx, column=3).value=data.get('Finalizado')
seleccionador_opciones=input('Selecciona una de las anteriores: ')
for x in opciones_usuario.keys():
    if seleccionador_opciones==x:        
        if seleccionador_opciones=='0':
            print(read())
        elif seleccionador_opciones=='1':
            datos_dictionary:dict = {
                'Id': input('Escribe un Id: '), 
                'actividad': input('Que actividad realizaras: '), 
                'Estado': input('Estado de la actividad: '), 
                'Finalizado': input('Tiempo que terminara la actividad: ')}
            create_dato(datos_dictionary
        )
        elif seleccionador_opciones=='2':
            eliminar_dato=input('Escriba el id para eliminar el dato')
            for x in read():
                print(x)
        elif seleccionador_opciones=='3':
            Dato_a_traer=input('Escriba el "id" del dato: ')
            Lectura=read()
            for x in Lectura:
                id=str(x['id'])
                if id == Dato_a_traer:
                    print(x)
        elif seleccionador_opciones=='4':
            updates({'actividad':'Correr', 'Finalizado': 5})