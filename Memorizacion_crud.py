from openpyxl import load_workbook
archivo='./Guardar_cronograma.xlsx'
file=load_workbook(archivo)
def read():
    datos=[]
    sheet=file['Hoja1']
    too=list(sheet.rows)
    for x in too:
        dictionary={
            'id':x[0].value,
            'actividad':x[1].value,
            'Estado':x[2].value,
            'Finalizado':x[3].value
        }
        datos.append(dictionary)
    print(datos)
def create_dato(Valores:dict):
    sheet=file['Hoja1']
    id_row=sheet.max_row+1
    sheet.cell(column=1, row=id_row).value=Valores.get('id')
    sheet.cell(column=2, row=id_row).value=Valores.get('actividad')
    sheet.cell(column=3, row=id_row).value=Valores.get('Estado')
    sheet.cell(column=4, row=id_row).value=Valores.get('Finalizado')
    file.save(archivo)
    dict_valores={
    'id':int(input('Escribe un numero id: ')),
    'actividad':input('Esribe la actividad que realizaras: '),
    'Estado':input('Estado de la actividad: '),
    'Finalizado':input('Tiempo a terminar la tarea: ')}
def delete(id:int):
    sheet=file['Hoja1']
    for x in sheet.rows:
        if x[0].value==id:
            idx=x[0].row
            sheet.delete_rows(idx)
    file.save(archivo)
def updates(data_updates:dict, id:int):
    sheet=file['Hoja1']
    for x in sheet.rows:
        if x[0].value==id:
            idx=x[0].row
            if (data_updates.get('actividad')):sheet.cell(row=idx, column=2).value=data_updates.get('actividad')
            if (data_updates.get('Estado')):sheet.cell(column=3, row=idx).value=data_updates.get('Estado')
            if (data_updates.get('Finalizado')):sheet.cell(column=4, row=idx).value=data_updates.get('Finalizado')
    file.save(archivo)
    read()







def actualizar(id:int, datos:dict):
    sheet=file['Hoja1']
    for x in sheet.rows:
        if x[0].value==id:
            idx=x[0].row
            if(datos.get('actividad')):sheet.cell(row=idx, column=2).value=datos.get('actividad')
            if datos.get('Estado'):sheet.cell(row=idx, column=3).value=datos.get('Estado')
            if datos.get('Finalizado'):sheet.cell(row=idx, column=4).value=datos.get('Finalizado')
actualizar(5, {'Estado':'Finalizado'})

def read_file():
    todos=[]
    sheet=file['Hoja1']
    rows=list(sheet.rows)
    for row in rows:
        dictionary={
            'id':row[0].value,
            'actividad':row[1].value,
            'Estado':row[2].value,
            'Finalizado':row[3].value
        }
        todos.append(dictionary)
    print(todos)
def agregar(datos:dict):
    sheet=file['Hoja1']
    row_max=sheet.max_row+1
    sheet.cell(row=row_max, column=1).value=datos.get('Id')
    sheet.cell(row=row_max, column=2).value=datos.get('actividad')
    sheet.cell(row=row_max, column=3).value=datos.get('Estado')
    sheet.cell(column=4, row=row_max).value=datos.get('Finalizado')
    file.save(archivo)
def delete(id:int):
    sheet=file['Hoja1']
    for row in sheet.rows:
        if row[0].value==id:
            posicition=row[0].row
            sheet.delete_rows(posicition)
    file.save(archivo)
delete(1)